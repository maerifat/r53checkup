import sys
from termcolor import colored,cprint 
from time import time, sleep
import webbrowser
from boto3.session import Session
import argparse
import re
import openpyxl
from openpyxl.styles import Font,PatternFill
import ipaddress
import dns.resolver #pip3 install dnspython

def main():
    def is_ip(text):
        try:
            ipaddress.IPv4Address(text)  
            return True
        except ipaddress.AddressValueError:
            try:
                ipaddress.IPv6Address(text) 
                return True
            except ipaddress.AddressValueError:
                return False

    def print_event(eventmsg,color,on_color=None):
        if args.verbose:
            if not args.no_color:
                if on_color:
                    eventmsg=colored(eventmsg,color,on_color)
                else:
                    eventmsg=colored(eventmsg,color)
            print(eventmsg)

    def comma_separated_values(values):
        return values.split(',')

    #Argument parsing
    parser = argparse.ArgumentParser(description='Route53 Record Collector')

    parser.add_argument(
        '-u',
        '--start-url',
        metavar='Start URL',
        type=str,
        required=True,
        help='aws SSO start URL. eg. https://d-1010ad440.awsapps.com/start'
    )

    parser.add_argument(
        '-a',
        '--accounts',
        metavar='account_id',
        type=comma_separated_values,
        help='multiple account_ids separated with comma. eg. 122389992,31313313,31313133'
    )

    parser.add_argument(
        '-r',
        '--region',
        metavar='region_name',
        type=str,
        help='Region name. eg. us-east-1'
    )

    parser.add_argument(
        '-v',
        '--verbose',
        action='store_true',
        help='Enable verbose to get details.'
    )

    parser.add_argument(
        '-l',
        '--list',
        action='store_true',
        help='List all subdomains, without any verbose.'
    )

    parser.add_argument(
        '-nc',
        '--no-color',
        action='store_true',
        help='Color less standard output.'
    )

    parser.add_argument(
        '-t',
        '--types',
        metavar='record_type',
        type=comma_separated_values,
        help='DNS record types separated with comma. eg. a,cname'
    )

    parser.add_argument(
        '-e',
        '--exclude',
        metavar='regular_expression',
        type=str,
        help='exclude subdomains which match this regular expression. eg. ".*_domainkey.*"'
    )

    parser.add_argument(
        '-cd',
        '--check-dangling',
        action='store_true',
        help='Checks if the DNS record is dangling.'
    )

    parser.add_argument(
        '-o',
        '--output',
        metavar='file_name',
        type=str,
        help='File name to save as, file type is recognised from the extension. eg subdomains.xlsx'
    )

    args = parser.parse_args()

    def is_text():
        if args.output:
            filelocation=args.output
            if filelocation.endswith('txt'):
                return True

    def is_excel():
        if args.output:
            filelocation=args.output
            if filelocation.endswith(('xls','xlsx')):
                return True

    def file_location():
        if args.output:
            filelocation=args.output
            return filelocation

    if is_excel():
        workbook = openpyxl.Workbook()
        sheet=workbook.active
        if args.check_dangling:
            sheet_headers= ['Account_Id', 'Zone_Name','Record_Name', 'Record_Type','Is_Alias','Is_Dangling','Record_Value']
        else:
            sheet_headers= ['Account_Id', 'Zone_Name','Record_Name', 'Record_Type','Is_Alias','Record_Value']
        sheet.append(sheet_headers)
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 70

    def get_dns_value():
        global dns_value
        global is_alias
        if record.get('ResourceRecords'):
            dns_value=[value['Value'] for value in  record['ResourceRecords'] ]
            dns_value=" , ".join(dns_value)
            is_alias=False
        elif record.get('AliasTarget'):
            if record['AliasTarget'].get('DNSName'):
                dns_value=record['AliasTarget']['DNSName']
                is_alias=True     
            else:
                dns_value="dnsvalueerror1"
        else:
            dns_value="dnsvalueerror2"
        return dns_value

    def is_dangling(dns_value):
        if record['Type'] == 'CNAME' or is_alias:
            try:
                result = dns.resolver.resolve(dns_value)
                if result:
                    return "No"
            except dns.resolver.NXDOMAIN:
                return "Yes"
            except dns.resolver.Timeout:
                return "Time Out"
            except dns.resolver.NoAnswer:
                return "No Answwer"
        else: 
            return "NA"

    def append_row_to_sheet():
        if args.check_dangling:
            sheet_row=[account_id,zone_name,record['Name'].rstrip('.'),record['Type'],is_alias,is_dangling(dns_value),get_dns_value()]
        else:
            sheet_row=[account_id,zone_name,record['Name'].rstrip('.'),record['Type'],is_alias,get_dns_value()]
        sheet.append(sheet_row)

    if args.list:
        args.verbose=None

    session =Session()

    ###Skeleton Creation###

    #Input details
    start_url = args.start_url

    if args.region:
        region = args.region
    else:
        region = 'us-east-1'

    #OIDC Connection
    sso_oidc = session.client('sso-oidc')
    client_creds = sso_oidc.register_client(
        clientName='r53collector',
        clientType='public',
    )
    if client_creds:
        print_event("[+] Client credentials fetched Succussfully.","yellow")

    #Device Authorization initiation
    device_authorization = sso_oidc.start_device_authorization(
        clientId=client_creds['clientId'],
        clientSecret=client_creds['clientSecret'],
        startUrl=start_url,
    )

    if device_authorization:
        print_event("[+] Device authorization has been initiated through browser. Waiting for authorization...","yellow")

    #Browser Authorization 
    url = device_authorization['verificationUriComplete']
    device_code = device_authorization['deviceCode']
    expires_in = device_authorization['expiresIn']
    interval = device_authorization['interval']
    webbrowser.open(url, autoraise=True)

    #Function to iterate and check if authorization is complete
    def authwait():
        for n in range(1, expires_in // (interval+5) + 1):
            sleep(interval+5)
            try:
                global token
                token = sso_oidc.create_token(
                    grantType='urn:ietf:params:oauth:grant-type:device_code',
                    deviceCode=device_code,
                    clientId=client_creds['clientId'],
                    clientSecret=client_creds['clientSecret'],
                )
                if n>1:
                    print_event("\r[+] Device yet to be authorized in browser, waiting...","yellow")
                    print_event(f"[+] Authorization Successful after {n} attemps.","green")
        
                else:
                    print_event(f"[+] Authorization Successful in first attempt.","green")


                break
            except sso_oidc.exceptions.AuthorizationPendingException:
                if args.verbose:
                    if n==1:
                        cprint("Device yet to be authorized in browser, waiting...","red",attrs=["blink"],end='', flush=True)
                    else:
                        cprint("\rDevice yet to be authorized in browser, waiting...","red",attrs=["blink"],end='', flush=True)
                pass

    #Wait until authorization
    authwait()

    access_token = token['accessToken']
    sso = session.client('sso')
    account_list_raw = sso.list_accounts(
        accessToken=access_token,
        maxResults=1000  
    )
    ####Skeleton Completed####

    #Fetch all accessible accounts otherwise give list
    if args.accounts:
        account_list= args.accounts
    else:
        account_list =  [account['accountId'] for account in account_list_raw['accountList']]

    print_event(f'[+] Total accounts: {len(account_list)}','yellow')
    print_event(f"    {account_list}\n\n","cyan")

    combined_subdomains = set()

    def get_subdomains(zone_id):
        subdomains= []
        try:
            paginate_resource_record_sets = route53.get_paginator('list_resource_record_sets')
            for record_response in paginate_resource_record_sets.paginate(
                HostedZoneId = zone_id,
                MaxItems='300'
                ):

                global record
                for record in record_response['ResourceRecordSets']:

                    #if record['Type']  in ['SOA', 'NS', 'MX', 'TXT'] and not record['Name'].startswith('_'):
                    if args.types and not args.exclude:
                        dns_types = list(map(str.upper, args.types))
                        if record['Type'] in dns_types:
                            get_dns_value()
                            if is_excel():
                                append_row_to_sheet()
                            subdomains.append(record['Name'].rstrip('.'))
                            combined_subdomains.add(record['Name'].rstrip('.'))
                            print_event(f"{record['Type']} : {record['Name']} ==> {get_dns_value()}","magenta")

                    elif args.exclude and not args.types:
                        regex_pattern = args.exclude
                        if not re.match(regex_pattern, record['Name'].rstrip('.')):
                            get_dns_value()
                            if is_excel():
                                append_row_to_sheet()
                            subdomains.append(record['Name'].rstrip('.'))
                            combined_subdomains.add(record['Name'].rstrip('.'))
                            print_event(f"{record['Type']} : {record['Name']} ==> {get_dns_value()}","magenta")

                    elif args.types and args.exclude:
                        dns_types = list(map(str.upper, args.types))
                        regex_pattern = args.exclude
                        if (record['Type'] in dns_types) and (not re.match(regex_pattern, record['Name'].rstrip('.'))):
                            get_dns_value()
                            if is_excel():
                                append_row_to_sheet()
                            subdomains.append(record['Name'].rstrip('.'))
                            combined_subdomains.add(record['Name'].rstrip('.'))
                            print_event(f"{record['Type']} : {record['Name']} ==> {get_dns_value()}","magenta")              
                    else:
                        get_dns_value()
                        if is_excel():
                            append_row_to_sheet()
                        subdomains.append(record['Name'].rstrip('.'))
                        combined_subdomains.add(record['Name'].rstrip('.'))
                        print_event(f"{record['Type']} : {record['Name']} ==> {get_dns_value()}","magenta")
        except Exception as e:
            print(f"Failed to get subdomains for zone {zone_id}: {e}")

        return subdomains
    #Iterate through accounts

    for account_id in account_list:
        account_roles = sso.list_account_roles(
            accessToken=access_token,
            accountId=account_id
        )
        roleNames = [role['roleName'] for role in account_roles['roleList']]
        
        #print(roleNames)

        try: 
    #Get credentials for each account
            role_creds = sso.get_role_credentials(
                roleName='Security_Audit',
                accountId=account_id,
                accessToken=access_token,
            )

            #Create session with these credentials
            session = Session(
                region_name=region,
                aws_access_key_id=role_creds['roleCredentials']['accessKeyId'],
                aws_secret_access_key=role_creds['roleCredentials']['secretAccessKey'],
                aws_session_token=role_creds['roleCredentials']['sessionToken'],
            )

            #Route53 client
            route53 = session.client('route53')      
            print_event(f"[+] Route53 DNS records in account {account_id}:","yellow","on_blue")

            paginate_hosted_zones = route53.get_paginator('list_hosted_zones')
            for zone_response in paginate_hosted_zones.paginate():
                for zone in zone_response['HostedZones']:
                    zone_id = zone['Id']
                    zone_name = zone['Name'].rstrip('.')
                    print_event("","green")
                    print_event(f"[+] Route53 DNS records in ZoneName {zone_name}:","yellow","on_light_blue")
                    print_event("","green")
                    subdomains = get_subdomains(zone_id)
            print_event("","green")
            print_event("","green")
        except:
            cprint(f"You do not have enough privileges in account {account_id}!", "red", attrs=["bold"], file=sys.stderr)

    print_event(f"[+] Unique subdomains across all accounts: {len(combined_subdomains)}","yellow","on_blue")
    for subdomain in combined_subdomains:
        print_event(f"    {subdomain}", "light_cyan")
        
    if is_text():
        with open(file_location(),'w') as textfile:
            for subdomain in combined_subdomains:
                textfile.write(subdomain+'\n')
        print_event(f"\n[+] All subdomains have been saved in text format in {file_location()}","yellow")

    if is_excel():

        header_font = Font(color="FFFFFF", bold=True)  # White bold font
        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Blue background

    # Apply font and fill to header row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=6):
            for cell in row:
                if cell.value == "Yes":
                    cell.font = Font(color="FF0000")  # Red font color
        workbook.save(file_location())
        print_event(f"\n[+] All data has been saved in excel format in {file_location()}","yellow")

if __name__=='__main__':
    main()
